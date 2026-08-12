"""Mass Opening Balance Import — service layer.

Property-level bulk import of opening balances for Tenants (payer type
`rental`) or Account Holders (payer type `levy`). The Property is the
operational aggregation unit; the individual account remains the accounting
unit. The engine:

  * generates a Property-specific Excel template (existing account codes only),
  * parses + validates an upload without ever creating an Account/Sub-Account,
  * enforces payer-type category rules, property membership, currency and
    duplicate checks,
  * posts valid rows through the accounting engine as proper opening-balance
    journals (offset to the 9000 Opening Balances account), and
  * reverses an entire batch while preserving structure and audit trail.

Sign convention (matches the tenant/account-holder pocket, which is
debit-normal): a POSITIVE amount is a debit (opening arrear the account owes);
a NEGATIVE amount is a credit (opening prepayment).
"""
from collections import defaultdict
from decimal import Decimal, InvalidOperation
import io

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from apps.masterfile.models import RentalTenant
from .account_coding import (
    template_categories, SUPPORTED_CURRENCIES, payer_side_label,
)
from .models import (
    OpeningBalanceImportBatch, OpeningBalanceImportRow,
    SubsidiaryAccount, SubsidiaryTransaction,
    ChartOfAccount, Journal, JournalEntry, AuditTrail,
)

# Category slug <-> human label used in the spreadsheet header.
CATEGORY_LABEL = {
    'rent': 'Rent', 'vat': 'VAT', 'deposit': 'Deposit',
    'levy': 'Levy', 'special_levy': 'Special Levy',
    'maintenance': 'Maintenance', 'parking': 'Parking', 'rates': 'Rates',
    'general': 'General',
}
LABEL_TO_CATEGORY = {v.lower(): k for k, v in CATEGORY_LABEL.items()}
# accept a few aliases
LABEL_TO_CATEGORY.update({'special levy': 'special_levy'})

OPENING_BALANCE_GL_CODE = '9000'


# ── helpers ──────────────────────────────────────────────────────────────
def payer_type_for(account_type):
    """RentalTenant.account_type ('rental'/'levy') for a batch AccountType."""
    return OpeningBalanceImportBatch.ACCOUNT_TYPE_TO_PAYER.get(account_type, 'rental')


def eligible_tenants(property_id, account_type):
    """Active RentalTenants of the matching payer type under a property.

    A tenant attaches to a property directly (its unit) or via a lease
    (denormalised property FK or the lease's unit). Mirrors the report-layer
    3-way OR so no eligible account is missed."""
    payer = payer_type_for(account_type)
    return (RentalTenant.objects.filter(is_active=True, account_type=payer)
            .filter(Q(unit__property_id=property_id) |
                    Q(leases__property_id=property_id) |
                    Q(leases__unit__property_id=property_id))
            .distinct())


def _category_columns(account_type):
    """Ordered category slugs that form the template's amount columns."""
    return template_categories(payer_type_for(account_type))


# ── template generation ──────────────────────────────────────────────────
def generate_template_xlsx(property_obj, account_type):
    """Return .xlsx bytes: one row per eligible account, one amount column per
    permitted category. The user fills amounts only — account codes, names and
    payer types come from the system so balances can't land on the wrong
    account. Add a second row for the same account with a different Currency to
    enter a foreign-currency opening balance."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    cats = _category_columns(account_type)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Opening Balances'

    header = ['Account Code', 'Account Name', 'Payer Type', 'Currency'] + \
             [CATEGORY_LABEL[c] for c in cats]
    ws.append(header)
    bold = Font(bold=True)
    fill = PatternFill('solid', fgColor='E8EEF7')
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.fill = fill

    payer_label = payer_side_label(payer_type_for(account_type))
    tenants = eligible_tenants(property_obj.id, account_type).order_by('code')
    for t in tenants:
        ws.append([t.code, t.name, payer_label, 'USD'] + [None] * len(cats))

    # Reasonable column widths.
    widths = [16, 32, 12, 10] + [14] * len(cats)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── parsing ──────────────────────────────────────────────────────────────
def _to_decimal(value):
    """Parse a spreadsheet cell into a signed Decimal, or None if not numeric."""
    if value is None:
        return None
    s = str(value).strip()
    if s == '':
        return None
    neg = False
    # Accept accounting-style parentheses for negatives, and thousands commas.
    if s.startswith('(') and s.endswith(')'):
        neg = True
        s = s[1:-1]
    s = s.replace(',', '').replace('$', '').strip()
    try:
        d = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    return -d if neg else d


def parse_upload(file_bytes):
    """Read a wide-format opening-balance sheet into flat records.

    Yields one record per non-zero amount cell:
      {source_row, account_code, account_name, currency, category,
       amount (Decimal|None), amount_ok (bool)}
    Zero / blank amounts are ignored (they neither post nor error — a zero
    opening balance must never create or delete a sub-account)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []

    col = {}
    cat_cols = {}  # column index -> category slug
    for idx, h in enumerate(header):
        if h is None:
            continue
        key = str(h).strip().lower()
        if key in ('account code', 'account_code', 'code'):
            col['code'] = idx
        elif key in ('account name', 'account_name', 'name'):
            col['name'] = idx
        elif key in ('currency', 'ccy'):
            col['currency'] = idx
        elif key in ('payer type', 'payer_type', 'account type', 'account_type'):
            col['payer'] = idx
        elif key in LABEL_TO_CATEGORY:
            cat_cols[idx] = LABEL_TO_CATEGORY[key]

    def cell(row, name):
        i = col.get(name)
        if i is None or i >= len(row):
            return None
        return row[i]

    records = []
    for r_i, row in enumerate(rows, start=2):  # header is row 1
        code = cell(row, 'code')
        if code is None or str(code).strip() == '':
            continue
        code = str(code).strip().upper()
        name = cell(row, 'name')
        name = str(name).strip() if name is not None else ''
        ccy = cell(row, 'currency')
        ccy = (str(ccy).strip().upper() if ccy is not None and str(ccy).strip() else 'USD')
        for c_idx, slug in cat_cols.items():
            if c_idx >= len(row):
                continue
            raw = row[c_idx]
            if raw is None or str(raw).strip() == '':
                continue
            amt = _to_decimal(raw)
            if amt is not None and amt == 0:
                continue  # explicit zero — ignore
            records.append({
                'source_row': r_i,
                'account_code': code,
                'account_name': name,
                'currency': ccy,
                'category': slug,
                'amount': amt,
                'amount_ok': amt is not None,
            })
    return records


# ── validation ───────────────────────────────────────────────────────────
def validate_records(batch, records):
    """Validate parsed records against the selected property/account type and
    the established account-structure rules. Creates OpeningBalanceImportRow
    rows (valid + error), fills batch counts/totals, and returns a summary.

    Never creates accounts or sub-accounts; a missing pocket is a row error."""
    payer = batch.payer_type()
    allowed = set(template_categories(payer))
    payer_label = payer_side_label(payer)

    codes = {r['account_code'] for r in records}
    # All matching accounts (incl. wrong type / other property) so we can give
    # a precise reason, and soft-deleted are surfaced as inactive.
    tmap = {t.code: t for t in RentalTenant.all_objects.filter(code__in=codes)}
    eligible_ids = set(eligible_tenants(batch.property_id, batch.account_type)
                       .values_list('id', flat=True))

    # Prefetch existing pockets for the referenced tenants (get-only structure).
    tenant_ids = [t.id for t in tmap.values()]
    pockets = {}
    for sa in SubsidiaryAccount.objects.filter(tenant_id__in=tenant_ids):
        pockets[(sa.tenant_id, sa.category)] = sa

    seen = set()  # (code, category, currency) duplicate guard
    totals = defaultdict(lambda: {'debit': Decimal('0'), 'credit': Decimal('0')})
    out_rows = []
    valid = 0

    for rec in records:
        code = rec['account_code']
        category = rec['category']
        currency = rec['currency']
        amount = rec['amount']
        row = OpeningBalanceImportRow(
            batch=batch, source_row=rec['source_row'], account_code=code,
            account_name=rec['account_name'], category=category,
            currency=currency, amount=amount or Decimal('0'),
        )

        error = None
        tenant = tmap.get(code)
        if not rec['amount_ok']:
            error = 'Amount is not a valid number'
        elif tenant is None:
            error = 'Account does not exist'
        elif tenant.account_type != payer:
            error = (f'Account is a {payer_side_label(tenant.account_type)} account, '
                     f'not a {payer_label} account')
        elif tenant.id not in eligible_ids:
            error = 'Account does not belong to the selected property'
        elif not tenant.is_active:
            error = 'Account is inactive'
        elif category not in allowed:
            error = f'{CATEGORY_LABEL.get(category, category)} is not permitted for a {payer_label} payer'
        elif currency not in SUPPORTED_CURRENCIES:
            error = f'Unsupported currency "{currency}"'
        elif (code, category, currency) in seen:
            error = 'Duplicate row for this account / category / currency'
        else:
            pocket = pockets.get((tenant.id, category))
            if pocket is None:
                # Never create structure — report and let the user fix it.
                error = f'Sub-account for {CATEGORY_LABEL.get(category, category)} does not exist'
            else:
                row.tenant = tenant
                row.subsidiary_account = pocket

        if error:
            row.status = OpeningBalanceImportRow.Status.ERROR
            row.error_message = error
        else:
            row.status = OpeningBalanceImportRow.Status.VALID
            seen.add((code, category, currency))
            if amount > 0:
                totals[currency]['debit'] += amount
            else:
                totals[currency]['credit'] += -amount
            valid += 1
        out_rows.append(row)

    OpeningBalanceImportRow.objects.bulk_create(out_rows)

    batch.record_count = len(out_rows)
    batch.valid_count = valid
    batch.error_count = len(out_rows) - valid
    batch.totals = {
        ccy: {'debit': str(v['debit']), 'credit': str(v['credit'])}
        for ccy, v in totals.items()
    }
    batch.status = (OpeningBalanceImportBatch.Status.VALIDATED if valid
                    else OpeningBalanceImportBatch.Status.FAILED)
    batch.save()
    return summary(batch)


def summary(batch):
    return {
        'batch_number': batch.batch_number,
        'status': batch.status,
        'record_count': batch.record_count,
        'valid_count': batch.valid_count,
        'error_count': batch.error_count,
        'totals': batch.totals,
    }


# ── posting ──────────────────────────────────────────────────────────────
def _opening_balance_gl_account():
    acct, _ = ChartOfAccount.objects.get_or_create(
        code=OPENING_BALANCE_GL_CODE,
        defaults={'name': 'Opening Balances', 'account_type': 'equity',
                  'account_subtype': 'retained_earnings', 'is_system': True},
    )
    return acct


@transaction.atomic
def post_batch(batch, user=None, allow_partial=False):
    """Post the batch's valid rows through the accounting engine.

    One Journal per currency: each valid row becomes a subsidiary-only line on
    its (payer, category) pocket (auto-mirrored to the 1300 control account),
    balanced by a single line to the 9000 Opening Balances account. Posting
    goes through Journal.post so the trial balance stays balanced, statements
    show the opening entry (source_type='opening_balance'), and the audit trail
    is written. Never creates accounts or sub-accounts."""
    if batch.status not in (OpeningBalanceImportBatch.Status.VALIDATED,
                            OpeningBalanceImportBatch.Status.PARTIALLY_FAILED):
        raise ValueError('Only a validated batch can be posted')
    if batch.error_count and not allow_partial:
        raise ValueError(
            'Batch has errors. Resolve them and re-upload, or explicitly choose '
            'to post the valid rows only.')

    valid_rows = list(batch.rows.filter(status=OpeningBalanceImportRow.Status.VALID)
                      .select_related('subsidiary_account'))
    if not valid_rows:
        raise ValueError('No valid rows to post')

    ob_account = _opening_balance_gl_account()
    by_ccy = defaultdict(list)
    for row in valid_rows:
        by_ccy[row.currency].append(row)

    journal_ids = []
    for ccy, rows in by_ccy.items():
        journal = Journal.objects.create(
            journal_type=Journal.JournalType.GENERAL,
            date=batch.date,
            description=f'Opening balances {batch.batch_number} — {batch.property.name}',
            reference=batch.batch_number,
            currency=ccy,
            created_by=user,
        )
        net = Decimal('0')
        row_entries = []
        for row in rows:
            amt = row.amount
            je = JournalEntry.objects.create(
                journal=journal,
                subsidiary_account=row.subsidiary_account,
                description=f'Opening balance {row.account_code} · {CATEGORY_LABEL.get(row.category, row.category)}',
                debit_amount=amt if amt > 0 else None,
                credit_amount=(-amt) if amt < 0 else None,
                source_type='opening_balance',
                source_id=batch.id,
            )
            row_entries.append((row, je))
            net += amt

        # Single balancing line to the Opening Balances (9000) account.
        if net > 0:
            JournalEntry.objects.create(
                journal=journal, account=ob_account,
                description=f'Opening balances offset {batch.batch_number}',
                credit_amount=net,
            )
        elif net < 0:
            JournalEntry.objects.create(
                journal=journal, account=ob_account,
                description=f'Opening balances offset {batch.batch_number}',
                debit_amount=-net,
            )
        # net == 0: pocket debits and credits already balance within the journal.

        journal.post(user)
        journal_ids.append(journal.id)

        # Link each row to the subsidiary transaction it produced.
        for row, je in row_entries:
            st = SubsidiaryTransaction.objects.filter(journal_entry=je).first()
            row.subsidiary_transaction = st
            row.status = OpeningBalanceImportRow.Status.POSTED
            row.save(update_fields=['subsidiary_transaction', 'status'])

    batch.posted_journal_ids = journal_ids
    batch.posted_by = user
    batch.posted_at = timezone.now()
    batch.status = (OpeningBalanceImportBatch.Status.PARTIALLY_FAILED
                    if batch.error_count else OpeningBalanceImportBatch.Status.POSTED)
    batch.save()

    AuditTrail.objects.create(
        action='opening_balance_batch_posted',
        model_name='OpeningBalanceImportBatch',
        record_id=batch.id,
        changes={
            'batch_number': batch.batch_number,
            'property_id': batch.property_id,
            'account_type': batch.account_type,
            'posted_rows': len(valid_rows),
            'skipped_errors': batch.error_count,
            'journals': journal_ids,
            'totals': batch.totals,
        },
        user=user,
    )
    return batch


@transaction.atomic
def reverse_batch(batch, reason, user=None):
    """Reverse every journal this batch posted, restoring balances while keeping
    the audit trail. Accounts and sub-accounts are never deleted."""
    if batch.status not in (OpeningBalanceImportBatch.Status.POSTED,
                            OpeningBalanceImportBatch.Status.PARTIALLY_FAILED):
        raise ValueError('Only a posted batch can be reversed')
    if not reason:
        raise ValueError('A reversal reason is required')

    reversed_journals = []
    for jid in batch.posted_journal_ids or []:
        j = Journal.objects.filter(id=jid, status=Journal.Status.POSTED).first()
        if j:
            rev = j.reverse(reason, user)
            reversed_journals.append(rev.id)

    batch.status = OpeningBalanceImportBatch.Status.REVERSED
    batch.reversed_by = user
    batch.reversed_at = timezone.now()
    batch.save(update_fields=['status', 'reversed_by', 'reversed_at', 'updated_at'])

    AuditTrail.objects.create(
        action='opening_balance_batch_reversed',
        model_name='OpeningBalanceImportBatch',
        record_id=batch.id,
        changes={
            'batch_number': batch.batch_number,
            'reason': reason,
            'reversal_journals': reversed_journals,
        },
        user=user,
    )
    return batch


# ── error report ─────────────────────────────────────────────────────────
def error_report_xlsx(batch):
    """.xlsx of every error row (source row, account, category, currency,
    amount, problem) so the user can correct the file and re-upload."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Errors'
    header = ['Row', 'Account Code', 'Category', 'Currency', 'Amount', 'Problem']
    ws.append(header)
    bold = Font(bold=True)
    fill = PatternFill('solid', fgColor='F7E8E8')
    for col in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col)
        c.font = bold
        c.fill = fill

    for row in batch.rows.filter(status=OpeningBalanceImportRow.Status.ERROR):
        ws.append([row.source_row, row.account_code,
                   CATEGORY_LABEL.get(row.category, row.category),
                   row.currency, float(row.amount), row.error_message])

    widths = [8, 16, 16, 10, 14, 48]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
